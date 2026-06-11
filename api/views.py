from django.contrib.auth import get_user_model
User = get_user_model()
from django.db.models import Sum, Q, Count
from django.utils import timezone
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
import random
import string
from rest_framework import generics, permissions
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework import viewsets
from rest_framework.exceptions import PermissionDenied
from rest_framework.views import APIView
from .models import Offence, Offender, Booking, User, Vehicle, DriverInformation, SMSLog
from .serializers import OffenceSerializer, OffenderSerializer, BookingSerializer, RegisterSerializer, CitizenRegisterSerializer, CustomTokenObtainPairSerializer, VehicleSerializer, PaymentSerializer, DriverInformationSerializer, SMSLogSerializer
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
import os
import requests
import hmac
import hashlib
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db import transaction
from rest_framework.filters import SearchFilter  # Phase 3: SMS API
import csv  # Phase 3: Bulk import
from io import StringIO  # Phase 3: Bulk import
from django.core.exceptions import ValidationError as DjangoValidationError  # Phase 3: Validation
from .tasks import retry_failed_sms

def generate_reference_id():
    # string.ascii_uppercase gives us 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # string.digits gives us '0123456789'
    # We combine them into one giant bucket of characters
    allowed_chars = string.ascii_uppercase + string.digits
    
    # random.choices picks 6 random characters from that bucket
    random_part = ''.join(random.choices(allowed_chars, k=6))
    
    # We return the final formatted string using an f-string
    return f"OYR-{random_part}"

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class IsAdminOrReadOnly(permissions.BasePermission):
    """
    Custom permission: Anyone can view the list of offences,
    but only Superusers (Admins) can add, edit, or delete them.
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS: # GET requests
            return True
        return request.user and request.user.is_superuser # POST, PUT, PATCH, DELETE

class OffenceViewSet(viewsets.ModelViewSet):
    
    from .models import Offence
    from .serializers import OffenceSerializer
    
    queryset = Offence.objects.all().order_by('name')
    serializer_class = OffenceSerializer
    permission_classes = [IsAdminOrReadOnly]

class OffenderViewSet(viewsets.ModelViewSet):
    queryset = Offender.objects.all()
    serializer_class = OffenderSerializer

class BookingViewSet(viewsets.ModelViewSet):
    serializer_class = BookingSerializer

   
    def get_queryset(self):
        user = self.request.user
        
        if user.is_staff or user.is_superuser:
            # Show the master list of all tickets across the state
            return Booking.objects.all().order_by('-date_time')
        # Offenders are linked to User accounts by driver_license_number == User.username
        return Booking.objects.filter(offender__driver_license_number=user.username).order_by('-date_time')
            
    def perform_update(self, serializer):
        user = self.request.user
        ticket = serializer.instance
        
        # If someone is trying to change the payment_status...
        if 'payment_status' in self.request.data:
            new_status = self.request.data.get('payment_status')
            # ...and they are NOT an officer/admin
            if not user.is_staff and not user.is_superuser:
                # Allow the offender themselves to mark their ticket as Paid (so citizens can pay via frontend)
                # Offender is linked by driver_license_number == user.username
                if not (ticket.offender and ticket.offender.driver_license_number == user.username and new_status == 'Paid'):
                    # Kick them out with a 403 Forbidden error!
                    raise PermissionDenied("Security Alert: Citizens cannot manually change ticket payment statuses.")
            if user.is_staff and not user.is_superuser:
                if ticket.officer != user:
                    raise PermissionDenied("Accountability Alert: You can only clear tickets that you personally issued.")
        
        # Otherwise, save the update normally
        serializer.save()

        # If the booking was just marked Paid, create a Payment record if one doesn't exist
        try:
            from .models import Payment
            if 'payment_status' in self.request.data and self.request.data.get('payment_status') == 'Paid':
                # create a payment record if none exists for this booking
                if not Payment.objects.filter(booking=ticket).exists():
                    Payment.objects.create(booking=ticket, amount=ticket.amount_due, transaction_id=self.request.data.get('transaction_id'))
        except Exception:
            # If anything goes wrong here, don't block the update; just log to stdout for now
            import traceback
            traceback.print_exc()
  
    def perform_create(self, serializer):
        
        new_ref_id = generate_reference_id()
        selected_offence = serializer.validated_data.get('offence')
        calculated_fine = selected_offence.fine_amount
        
       
        serializer.save(
            reference_id=new_ref_id, 
            amount_due=calculated_fine,
            officer=self.request.user  
        )

    @action(detail=True, methods=['post'], url_path='pay')
    def pay(self, request, pk=None):
        booking = self.get_object()

        # Only allow officers or admins to register payments via this endpoint
        user = request.user
        if not user.is_staff and not user.is_superuser and not user.is_officer:
            return Response({'detail': 'Only officers or admins can register payments.'}, status=status.HTTP_403_FORBIDDEN)

        amount = booking.amount_due

        # Create a Payment record (transaction_id would come from a payment gateway)
        payment_data = {
            'booking': booking.id,
            'amount': amount,
            'transaction_id': request.data.get('transaction_id', None)
        }
        serializer = PaymentSerializer(data=payment_data)
        if serializer.is_valid():
            serializer.save()
            booking.payment_status = 'Paid'
            booking.save(update_fields=['payment_status'])
            return Response({'detail': 'Payment recorded and booking marked as Paid.'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='verify-payment')
    def verify_payment(self, request, pk=None):
        """Verify a payment reference with Paystack and mark booking paid if valid.

        Body: { "reference": "PSK_REF_123" }
        """
        booking = self.get_object()
        ref = request.data.get('reference')
        print(f"\n🔍 [verify_payment] Called for booking {pk} with reference: {ref}")
        
        if not ref:
            print(f"❌ [verify_payment] No reference provided")
            return Response({'detail': 'reference is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Support multiple possible .env key names (some deployments use *_KEY suffix)
        paystack_secret = os.getenv('PAYSTACK_SECRET') or os.getenv('PAYSTACK_SECRET_KEY') or os.getenv('PAYSTACK_SK')
        if not paystack_secret:
            print(f"❌ [verify_payment] Paystack secret not configured")
            return Response({'detail': 'Paystack secret not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Call Paystack verify endpoint
        print(f"📤 [verify_payment] Calling Paystack API for reference: {ref}")
        try:
            resp = requests.get(f'https://api.paystack.co/transaction/verify/{ref}', headers={'Authorization': f'Bearer {paystack_secret}'}, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            print(f"✅ [verify_payment] Paystack response: {data}")
        except Exception as e:
            print(f"❌ [verify_payment] Paystack API failed: {str(e)}")
            return Response({'detail': f'Failed to verify with Paystack: {str(e)}'}, status=status.HTTP_502_BAD_GATEWAY)

        if not data.get('status') or not data.get('data'):
            print(f"❌ [verify_payment] Invalid response from Paystack: {data}")
            return Response({'detail': 'Invalid response from Paystack'}, status=status.HTTP_502_BAD_GATEWAY)

        pay_data = data['data']
        # Paystack returns amount in kobo (lowest currency unit)
        paid_amount = pay_data.get('amount') / 100.0 if pay_data.get('amount') is not None else None
        status_str = pay_data.get('status')
        print(f"💰 [verify_payment] Paystack status: {status_str}, amount: {paid_amount}")

        if status_str != 'success':
            print(f"❌ [verify_payment] Payment not successful: {status_str}")
            return Response({'detail': f'Payment not successful: {status_str}'}, status=status.HTTP_400_BAD_REQUEST)

        # Optional: ensure amount matches
        if paid_amount is not None and float(paid_amount) < float(booking.amount_due):
            print(f"❌ [verify_payment] Paid amount ({paid_amount}) < booking amount ({booking.amount_due})")
            return Response({'detail': 'Paid amount is less than booking amount'}, status=status.HTTP_400_BAD_REQUEST)

        # Record payment and mark booking paid
        from .models import Payment
        print(f"💾 [verify_payment] Updating booking {booking.id} to Paid status...")
        # Ensure the payment creation and booking update are atomic to avoid race conditions
        with transaction.atomic():
            payment, created = Payment.objects.get_or_create(booking=booking, transaction_id=ref, defaults={'amount': booking.amount_due})
            booking.payment_status = 'Paid'
            booking.save(update_fields=['payment_status'])
            print(f"✅ [verify_payment] Payment record {'created' if created else 'updated'}: {payment.id}")
            print(f"✅ [verify_payment] Booking {booking.id} marked as Paid")
        return Response({'detail': 'Payment verified and booking marked as Paid.'})
    
    @action(detail=True, methods=['post'], url_path='initialize-payment')
    def initialize_payment(self, request, pk=None):
        booking = self.get_object()
        
        # Talk to Paystack securely from the server
        url = "https://api.paystack.co/transaction/initialize"
        headers = {
            "Authorization": f"Bearer {settings.PAYSTACK_SECRET_KEY}", # Uses your secret key!
            "Content-Type": "application/json"
        }
        
        # Create a unique reference and set the return URL
        reference = f"OYRTMA-{booking.id}-{int(timezone.now().timestamp())}"
        
        data = {
            "email": request.user.email or "driver@oyrtma.gov.ng",
            "amount": int(float(booking.amount_due) * 100),
            "reference": reference,
            # This is where Paystack sends them AFTER they pay:
            "callback_url": "http://localhost:5173/payment-callback" 
        }
        
        response = requests.post(url, headers=headers, json=data)
        
        if response.status_code == 200:
            # Send the secure checkout URL back to React
            return Response(response.json())
        else:
            return Response({"error": "Paystack server error"}, status=400)


# Paystack webhook receiver
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def paystack_webhook(request):
    """Handle Paystack webhooks. Validates signature and updates booking/payment state.

    Paystack sends x-paystack-signature header which is HMAC SHA512 of the request.body using PAYSTACK_SECRET.
    """
    # Support multiple possible .env key names
    paystack_secret = os.getenv('PAYSTACK_SECRET') or os.getenv('PAYSTACK_SECRET_KEY') or os.getenv('PAYSTACK_SK')
    if not paystack_secret:
        return Response({'detail': 'Paystack secret not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    signature = request.headers.get('x-paystack-signature')
    raw_body = request.body
    computed = hmac.new(paystack_secret.encode('utf-8'), raw_body, hashlib.sha512).hexdigest()
    if not hmac.compare_digest(computed, signature or ''):
        return Response({'detail': 'Invalid signature'}, status=status.HTTP_400_BAD_REQUEST)

    payload = request.data
    # handle payment.success event
    event = payload.get('event')
    data = payload.get('data', {})
    if data and data.get('status') == 'success':
        reference = data.get('reference')
        # verify server-side similarly to ensure idempotency
        try:
            resp = requests.get(f'https://api.paystack.co/transaction/verify/{reference}', headers={'Authorization': f'Bearer {paystack_secret}'}, timeout=10)
            resp.raise_for_status()
            result = resp.json()
            if result.get('status') and result['data'].get('status') == 'success':
                # Try to find booking by reference id present in metadata or map via transaction reference
                # If you previously stored transaction_id on Payment, use that. We'll try booking with matching reference_id first.
                reference_id = data.get('metadata', {}).get('reference_id') or None
                from .models import Booking, Payment
                booking = None
                if reference_id:
                    booking = Booking.objects.filter(reference_id=reference_id).first()
                if not booking:
                    # try to find payment with transaction_id
                    payment = Payment.objects.filter(transaction_id=reference).first()
                    if payment:
                        booking = payment.booking

                if booking:
                    # Use atomic block to ensure idempotent safe write
                    from .models import Payment
                    with transaction.atomic():
                        Payment.objects.get_or_create(booking=booking, transaction_id=reference, defaults={'amount': booking.amount_due})
                        booking.payment_status = 'Paid'
                        booking.save(update_fields=['payment_status'])
        except Exception:
            pass

    return Response({'status': 'ok'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_paystack_config(request):
    """Return public Paystack configuration for frontend (public key)."""
    # Support multiple possible .env key names for public key
    public = os.getenv('PAYSTACK_PUBLIC') or os.getenv('PAYSTACK_PUBLIC_KEY') or os.getenv('PAYSTACK_PK')
    if not public:
        return Response({'detail': 'Paystack public key not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return Response({'publicKey': public})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def verify_payment_by_reference(request):
    """Verify a Paystack transaction by reference and mark the related booking Paid.

    Body: { "reference": "PSK_REF_123" }
    This endpoint will:
      - call Paystack verify API
      - read metadata.reference_id from Paystack response
      - find booking by reference_id (or by existing Payment.transaction_id)
      - create Payment record and mark booking Paid
    """
    ref = request.data.get('reference')
    if not ref:
        return Response({'detail': 'reference is required'}, status=status.HTTP_400_BAD_REQUEST)

    paystack_secret = os.getenv('PAYSTACK_SECRET') or os.getenv('PAYSTACK_SECRET_KEY') or os.getenv('PAYSTACK_SK')
    if not paystack_secret:
        return Response({'detail': 'Paystack secret not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        resp = requests.get(f'https://api.paystack.co/transaction/verify/{ref}', headers={'Authorization': f'Bearer {paystack_secret}'}, timeout=10)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        return Response({'detail': f'Failed to verify with Paystack: {str(e)}'}, status=status.HTTP_502_BAD_GATEWAY)

    if not data.get('status') or not data.get('data'):
        return Response({'detail': 'Invalid response from Paystack'}, status=status.HTTP_502_BAD_GATEWAY)

    pay_data = data['data']
    status_str = pay_data.get('status')
    paid_amount = pay_data.get('amount') / 100.0 if pay_data.get('amount') is not None else None

    if status_str != 'success':
        return Response({'detail': f'Payment not successful: {status_str}'}, status=status.HTTP_400_BAD_REQUEST)

    # Try to locate booking using metadata.reference_id or existing Payment
    reference_id = pay_data.get('metadata', {}).get('reference_id') or None
    from .models import Booking, Payment
    booking = None
    if reference_id:
        booking = Booking.objects.filter(reference_id=reference_id).first()

    if not booking:
        payment = Payment.objects.filter(transaction_id=ref).first()
        if payment:
            booking = payment.booking

    if not booking:
        return Response({'detail': 'Could not locate booking for this transaction reference'}, status=status.HTTP_404_NOT_FOUND)

    # Check amount
    if paid_amount is not None and float(paid_amount) < float(booking.amount_due):
        return Response({'detail': 'Paid amount is less than booking amount'}, status=status.HTTP_400_BAD_REQUEST)

    payment, created = Payment.objects.get_or_create(booking=booking, transaction_id=ref, defaults={'amount': booking.amount_due})
    with transaction.atomic():
        payment, created = Payment.objects.get_or_create(booking=booking, transaction_id=ref, defaults={'amount': booking.amount_due})
        booking.payment_status = 'Paid'
        booking.save(update_fields=['payment_status'])
    return Response({'detail': 'Payment verified and booking marked as Paid.'})

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (AllowAny,) 
    serializer_class = RegisterSerializer

    def post(self, request, *args, **kwargs):
        raise PermissionDenied("Self-registration is disabled. Please contact the administrator.")



class CitizenRegisterView(generics.CreateAPIView):
    permission_classes = (AllowAny,) 
    serializer_class = CitizenRegisterSerializer

class AddVehicleView(generics.CreateAPIView):
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated] 

    def create(self, request, *args, **kwargs):
        plate_number = request.data.get('plate_number')
        vehicle_model = request.data.get('vehicle_model', '')

        if not plate_number:
            return Response(
                {"plate_number": ["This field is required."]},
                status=status.HTTP_400_BAD_REQUEST
            )

        plate_number = plate_number.strip().upper()

        # If vehicle_model was not supplied or empty, attempt to resolve it from the driver registry (DriverInformation)
        if not vehicle_model:
            try:
                registry_entry = DriverInformation.objects.get(plate_number=plate_number)
                if registry_entry.vehicle_type and registry_entry.vehicle_type != 'Unknown':
                    vehicle_model = registry_entry.vehicle_type
            except DriverInformation.DoesNotExist:
                pass

        try:
            current_driver = Offender.objects.get(driver_license_number=request.user.username)
        except Offender.DoesNotExist:
            return Response(
                {"detail": "Driver profile (Offender) not found for this user account."},
                status=status.HTTP_404_NOT_FOUND
            )

        # 1. Handle Driver Registry (DriverInformation)
        driver_info, di_created = DriverInformation.objects.get_or_create(
            plate_number=plate_number,
            defaults={
                'phone_number': current_driver.phone_number or '',
                'driver_name': current_driver.driver_name or 'Unknown',
                'state': 'Oyo',  # OYRTMA operates in Oyo State
                'license_number': current_driver.driver_license_number or '',
                'email': current_driver.email or '',
                'vehicle_type': vehicle_model or 'Unknown',
                'is_active': True
            }
        )
        if not di_created:
            # Sync existing driver information if missing details
            updated = False
            if not driver_info.is_active:
                driver_info.is_active = True
                updated = True
            if not driver_info.phone_number and current_driver.phone_number:
                driver_info.phone_number = current_driver.phone_number
                updated = True
            if not driver_info.driver_name or driver_info.driver_name == 'Unknown':
                if current_driver.driver_name:
                    driver_info.driver_name = current_driver.driver_name
                    updated = True
            if not driver_info.license_number and current_driver.driver_license_number:
                driver_info.license_number = current_driver.driver_license_number
                updated = True
            if not driver_info.email and current_driver.email:
                driver_info.email = current_driver.email
                updated = True
            if (not driver_info.vehicle_type or driver_info.vehicle_type == 'Unknown') and vehicle_model:
                driver_info.vehicle_type = vehicle_model
                updated = True
            if updated:
                driver_info.save()

        # 2. Handle Vehicle Record (assign/attach it to current driver)
        vehicle, v_created = Vehicle.objects.get_or_create(
            plate_number=plate_number,
            defaults={
                'owner': current_driver,
                'vehicle_model': vehicle_model
            }
        )
        if not v_created:
            vehicle.owner = current_driver
            if vehicle_model:
                vehicle.vehicle_model = vehicle_model
            vehicle.save()

        serializer = self.get_serializer(vehicle)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class VehiclesView(generics.ListAPIView):
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Look up the logged-in user and only return THEIR cars
        return Vehicle.objects.filter(owner__driver_license_number=self.request.user.username)


class AdminDashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Enforce strict RBAC: Only superusers/admins can access this
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only administrators can view state statistics.")

        # Calculate revenue metrics
        total_revenue = Booking.objects.filter(payment_status='Paid').aggregate(Sum('amount_due'))['amount_due__sum'] or 0
        pending_revenue = Booking.objects.filter(payment_status='Pending').aggregate(Sum('amount_due'))['amount_due__sum'] or 0
        
        # Calculate operational metrics
        total_tickets = Booking.objects.count()
        # Assuming officers are marked as 'is_staff' in your User model
        active_officers = User.objects.filter(is_staff=True, is_superuser=False).count()

        return Response({
            'total_revenue': total_revenue,
            'pending_revenue': pending_revenue,
            'total_tickets': total_tickets,
            'active_officers': active_officers
        })

class AdminOfficerManagementView(APIView):
    permission_classes = [IsAuthenticated]

    # --- 1. FETCH OFFICERS (Your perfectly updated code) ---
    def get(self, request, pk=None):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only admins can view officers.")
        
        User = get_user_model()
        from .models import Offender 
        
        citizen_usernames = Offender.objects.exclude(
            driver_license_number__isnull=True
        ).exclude(
            driver_license_number__exact=''
        ).values_list('driver_license_number', flat=True)
        
        users = User.objects.filter(
            Q(is_staff=True) | ~Q(username__in=citizen_usernames),
            is_superuser=False
        ).order_by('-date_joined').values(
            'id', 'username', 'email', 'first_name', 'last_name', 'is_staff', 'is_active', 'date_joined'
        )
        
        return Response(list(users))

    # --- 2. APPROVE OFFICER & UNLOCK ACCOUNT ---
    def patch(self, request, pk=None):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only admins can approve officers.")
        
        User = get_user_model()
        try:
            user = User.objects.get(pk=pk)
            
            # Grant the system badge
            user.is_staff = True  
            # Mark as official officer role
            user.is_officer = True 
            # Unlock the account (overriding the is_active=False from the serializer)
            user.is_active = True 
            
            user.save()

            if user.email:
                subject = "Your OYRTMA Officer Badge is Approved!"
                message = (
                    f"Hello {user.first_name},\n\n"
                    f"Your official OYRTMA Field Officer account has been fully approved and unlocked by Headquarters.\n\n"
                    f"You can now log into the Command Center and the Field Ticket App using your Staff ID: {user.username}\n\n"
                    f"Stay safe on the roads!\n"
                    f"- OYRTMA System Administrator"
                )
                
                try:
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER, # Sender
                        [user.email],             # Recipient
                        fail_silently=True        # Prevents server crash if email config is wrong
                    )
                except Exception as e:
                    print(f"Email failed to send: {e}")
            return Response({"message": "Officer approved and account unlocked successfully!"})
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

    # --- 3. REJECT / DELETE OFFICER ---
    def delete(self, request, pk=None):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only admins can remove officers.")
        
        User = get_user_model()
        try:
            user = User.objects.get(pk=pk)
            user.delete() 
            return Response({"message": "User removed successfully!"})
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

    # --- 4. ADMIN REGISTERS NEW OFFICER ---
    def post(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only admins can register new officers.")
        
        username = request.data.get('username')
        email = request.data.get('email')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        password = request.data.get('password')
        
        if not username or not email or not first_name or not last_name:
            return Response({"error": "Staff ID, Email, First Name, and Last Name are required."}, status=status.HTTP_400_BAD_REQUEST)
        
        User = get_user_model()
        if User.objects.filter(username=username).exists():
            return Response({"error": "This Staff ID Number is already registered in the system."}, status=status.HTTP_400_BAD_REQUEST)
            
        if User.objects.filter(email=email).exists():
            return Response({"error": "This Email Address is already registered."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Auto-generate password if not provided
        if not password:
            import secrets
            # Generate a clean 8-character password
            alphabet = string.ascii_letters + string.digits
            password = ''.join(secrets.choice(alphabet) for _ in range(8))
            
        try:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            user.is_staff = True
            user.is_officer = True
            user.is_active = True
            user.save()
            
            # Send Email
            subject = "Your OYRTMA Officer Credentials"
            message = (
                f"Hello {first_name} {last_name},\n\n"
                f"An official OYRTMA Field Officer account has been created for you by the System Administrator.\n\n"
                f"Your account details are:\n"
                f"Staff ID / Username: {username}\n"
                f"Password: {password}\n\n"
                f"You can log into the Command Center and the Field Ticket App using the following link:\n"
                f"http://localhost:5174/ \n\n"
                f"Please make sure to log in and keep your credentials secure.\n\n"
                f"Stay safe on the roads!\n"
                f"- OYRTMA System Administrator"
            )
            
            try:
                send_mail(
                    subject,
                    message,
                    settings.EMAIL_HOST_USER,
                    [email],
                    fail_silently=False
                )
                email_sent = True
            except Exception as e:
                print(f"Error sending email: {e}")
                email_sent = False
                
            return Response({
                "message": "Officer registered successfully!",
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "is_staff": user.is_staff,
                    "is_active": user.is_active,
                },
                "email_sent": email_sent,
                "generated_password": password
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({"error": f"Failed to register officer: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AdminDashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only admins can view stats.")

        User = get_user_model()
        
        # --- RESTORED: Month calculations needed for Revenue Growth ---
        now = timezone.now()
        first_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if first_this_month.month == 1:
            first_last_month = first_this_month.replace(year=first_this_month.year - 1, month=12)
        else:
            first_last_month = first_this_month.replace(month=first_this_month.month - 1)
        
        # Look for custom dates in the URL
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        # Base Querysets
        booking_query = Booking.objects.all()

        if start_date_str and end_date_str:
            # If dates provided, filter the main queries!
            booking_query = booking_query.filter(date_time__range=[start_date_str, end_date_str + " 23:59:59"])

        # 1. Base Stats 
        total_revenue = booking_query.filter(payment_status='Paid').aggregate(Sum('amount_due'))['amount_due__sum'] or 0
        pending_revenue = booking_query.filter(payment_status='Pending').aggregate(Sum('amount_due'))['amount_due__sum'] or 0
        total_tickets = booking_query.count()
        active_officers = User.objects.filter(is_staff=True, is_superuser=False).count()

        # 2. Analytics: Revenue Growth
        this_month_rev = booking_query.filter(payment_status='Paid', date_time__gte=first_this_month).aggregate(Sum('amount_due'))['amount_due__sum'] or 0
        last_month_rev = booking_query.filter(payment_status='Paid', date_time__gte=first_last_month, date_time__lt=first_this_month).aggregate(Sum('amount_due'))['amount_due__sum'] or 0

        revenue_growth = 0
        if last_month_rev > 0:
            revenue_growth = ((this_month_rev - last_month_rev) / last_month_rev) * 100
        elif this_month_rev > 0:
            revenue_growth = 100 

        # 3. Analytics: Top Crime Hotspots (Locations) - CHANGED TO booking_query
        top_locations = list(booking_query.values('location').annotate(count=Count('id')).order_by('-count')[:5])

        # 4. Analytics: Top Performing Officers (Catch Rate) - CHANGED TO booking_query
        top_officers = list(booking_query.exclude(officer__isnull=True).values(
            'officer__first_name', 'officer__last_name', 'officer__username'
        ).annotate(count=Count('id')).order_by('-count')[:5])

        # 5. Analytics: Most Frequent Offences - CHANGED TO booking_query
        top_offences = list(booking_query.values('offence__name').annotate(count=Count('id')).order_by('-count')[:5])

        return Response({
            "total_revenue": total_revenue,
            "pending_revenue": pending_revenue,
            "total_tickets": total_tickets,
            "active_officers": active_officers,
            "this_month_revenue": this_month_rev,
            "last_month_revenue": last_month_rev,
            "revenue_growth": round(revenue_growth, 1),
            "top_locations": top_locations,
            "top_officers": top_officers,
            "top_offences": top_offences
        })

# --- PASSWORD RESET REQUEST ---
class PasswordResetRequestView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        email = request.data.get('email')
        User = get_user_model()
        try:
            user = User.objects.get(email=email)
            # Generate a secure one-time token and encode the user's ID
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            
            reset_link = f"http://localhost:5173/reset-password/{uid}/{token}"
            
            # Send the email
            send_mail(
                "Password Reset Request - OYRTMA",
                f"Hello {user.first_name},\n\nYou requested to reset your password.\nClick the secure link below to create a new password:\n\n{reset_link}\n\nIf you did not request this, please ignore this email.",
                settings.EMAIL_HOST_USER,
                [user.email],
                fail_silently=False
            )
            return Response({"message": "A password reset link has been sent to your email."})
        except User.DoesNotExist:
            return Response({"error": "No user found with this email address."}, status=404)

# --- PASSWORD RESET CONFIRMATION ---
class PasswordResetConfirmView(APIView):
    permission_classes = [AllowAny]
    def post(self, request, uidb64, token):
        new_password = request.data.get('new_password')
        User = get_user_model()
        try:
            # Decode the user's ID
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
            
           
            if default_token_generator.check_token(user, token):
                user.set_password(new_password)
                user.save()
                return Response({"message": "Password successfully reset! You can now log in."})
            else:
                return Response({"error": "This reset link is invalid or has expired."}, status=400)
        except Exception:
            return Response({"error": "Invalid reset link."}, status=400)

class CitizenPasswordResetVerifyView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        license_number = request.data.get('driver_license')
        phone_number = request.data.get('phone_number')

        if not license_number or not phone_number:
            return Response({"error": "Both License Number and Phone Number are required."}, status=400)

        User = get_user_model()
        from .models import Offender 

        try:
            # Citizens log in with their license as their username
            user = User.objects.get(username=license_number)
            offender = Offender.objects.get(driver_license_number=license_number)

            # Security Check: Does the phone number match?
            if offender.phone_number == phone_number:
                # Match found! Generate the secure tokens
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                
                # Send the keys back to React to unlock the reset screen
                return Response({
                    "message": "Identity verified!",
                    "uidb64": uid,
                    "token": token
                })
            else:
                return Response({"error": "The Phone Number does not match our records for this License."}, status=400)
                
        except (User.DoesNotExist, Offender.DoesNotExist):
            return Response({"error": "No driver account found with this License Number."}, status=404)


# ========== SMS NOTIFICATION API ENDPOINTS (Phase 3) ==========

from rest_framework.filters import SearchFilter
from .models import DriverInformation, SMSLog
from .serializers import DriverInformationSerializer, SMSLogSerializer
import csv
from io import StringIO
from django.core.exceptions import ValidationError as DjangoValidationError


class IsOfficer(permissions.BasePermission):
    """
    Permission: Only officers and admins can access driver endpoints
    """
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and (
            request.user.is_staff or hasattr(request.user, 'officer')
        )


class IsAdmin(permissions.BasePermission):
    """
    Permission: Only admins/superusers can access
    """
    def has_permission(self, request, view):
        return request.user and request.user.is_staff


class DriverInformationViewSet(viewsets.ModelViewSet):
    """
    API endpoints for driver information management
    
    GET    /api/drivers/                 - List all drivers
    POST   /api/drivers/                 - Create driver
    GET    /api/drivers/{id}/            - Get driver details
    PUT    /api/drivers/{id}/            - Update driver
    DELETE /api/drivers/{id}/            - Delete driver
    POST   /api/drivers/bulk-import/     - Bulk import from CSV
    GET    /api/drivers/search/?plate=   - Search by plate number
    """
    
    queryset = DriverInformation.objects.all()
    serializer_class = DriverInformationSerializer
    permission_classes = [IsAuthenticated, IsOfficer]
    filter_backends = [SearchFilter]
    search_fields = ['plate_number', 'phone_number', 'driver_name', 'state']
    
    @action(detail=False, methods=['post'], url_path='bulk-import')
    def bulk_import(self, request):
        """
        Bulk import drivers from CSV file
        
        Expected CSV format:
        plate_number,phone_number,driver_name,state,license_number,email,vehicle_type
        
        Request: POST /api/drivers/bulk-import/ with file in 'file' field
        """
        try:
            file = request.FILES.get('file')
            if not file:
                return Response(
                    {'error': 'No file provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Read CSV file
            decoded_file = file.read().decode('utf-8')
            csv_data = csv.DictReader(StringIO(decoded_file))
            
            created_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(csv_data, start=2):  # Start from 2 (header is 1)
                try:
                    # Validate required fields
                    required_fields = ['plate_number', 'phone_number']
                    for field in required_fields:
                        if not row.get(field):
                            raise DjangoValidationError(f"Missing required field: {field}")
                    
                    # Format phone number
                    phone = row['phone_number'].strip()
                    if not phone.startswith('+234') and not phone.startswith('0'):
                        phone = '+234' + phone
                    elif phone.startswith('0'):
                        phone = '+234' + phone[1:]
                    elif not phone.startswith('+'):
                        phone = '+' + phone
                    
                    # Create or update driver
                    driver, created = DriverInformation.objects.update_or_create(
                        plate_number=row['plate_number'].strip(),
                        defaults={
                            'phone_number': phone,
                            'driver_name': row.get('driver_name', 'Unknown').strip(),
                            'state': row.get('state', 'Unknown').strip(),
                            'license_number': row.get('license_number', '').strip(),
                            'email': row.get('email', '').strip(),
                            'vehicle_type': row.get('vehicle_type', 'Unknown').strip(),
                            'is_active': True,
                        }
                    )
                    created_count += 1
                
                except Exception as e:
                    error_count += 1
                    errors.append({
                        'row': row_num,
                        'error': str(e),
                        'data': row
                    })
            
            return Response({
                'message': f'Bulk import completed',
                'created_or_updated': created_count,
                'errors': error_count,
                'error_details': errors[:10]  # Return first 10 errors
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Search drivers by plate number
        
        Usage: GET /api/drivers/search/?plate=ABC-123-XYZ
        """
        plate = request.query_params.get('plate')
        if not plate:
            return Response(
                {'error': 'plate parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            driver = DriverInformation.objects.get(plate_number=plate)
            serializer = self.get_serializer(driver)
            return Response(serializer.data)
        except DriverInformation.DoesNotExist:
            return Response(
                {'error': f'Driver with plate {plate} not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class SMSLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only API endpoints for SMS log management
    
    GET    /api/sms-logs/                         - List all SMS logs
    GET    /api/sms-logs/{id}/                    - Get SMS log details
    GET    /api/sms-logs/by-booking/{booking_id}/ - Get SMS logs for booking
    POST   /api/sms-logs/{id}/retry/              - Retry failed SMS
    """
    
    queryset = SMSLog.objects.all().order_by('-created_at')
    serializer_class = SMSLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filter_backends = [SearchFilter]
    search_fields = ['phone_number', 'booking__reference_id', 'status']
    
    @action(detail=True, methods=['post'])
    def retry(self, request, pk=None):
        """
        Manually retry sending a failed SMS
        
        Usage: POST /api/sms-logs/{id}/retry/
        """
        try:
            sms_log = self.get_object()
            
            if sms_log.status == 'sent':
                return Response(
                    {'error': 'Cannot retry an SMS that was already sent'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Queue retry task
            retry_failed_sms.delay(sms_log.id)
            
            return Response({
                'message': 'SMS retry task queued',
                'sms_log_id': sms_log.id,
                'task_status': 'queued'
            })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to queue retry task: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def by_booking(self, request):
        """
        Get all SMS logs for a specific booking
        
        Usage: GET /api/sms-logs/by-booking/?booking_id=123
        """
        booking_id = request.query_params.get('booking_id')
        if not booking_id:
            return Response(
                {'error': 'booking_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        sms_logs = self.queryset.filter(booking_id=booking_id)
        if not sms_logs.exists():
            return Response(
                {'error': f'No SMS logs found for booking {booking_id}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = self.get_serializer(sms_logs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get SMS statistics
        
        Usage: GET /api/sms-logs/stats/
        """
        from django.db.models import Count
        
        total = SMSLog.objects.count()
        sent = SMSLog.objects.filter(status='sent').count()
        failed = SMSLog.objects.filter(status='failed').count()
        pending = SMSLog.objects.filter(status='pending').count()
        
        return Response({
            'total': total,
            'sent': sent,
            'failed': failed,
            'pending': pending,
            'success_rate': f"{(sent/total*100):.1f}%" if total > 0 else "0%"
        })


from django.utils import timezone
from rest_framework.exceptions import PermissionDenied
from .models import TicketDispute
from .serializers import TicketDisputeSerializer, CitizenProfileSerializer, ChangePasswordSerializer

class TicketDisputeViewSet(viewsets.ModelViewSet):
    serializer_class = TicketDisputeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return TicketDispute.objects.filter(offender__driver_license_number=self.request.user.username).order_by('-submitted_at')

    def perform_create(self, serializer):
        booking = serializer.validated_data.get('booking')
        try:
            offender = Offender.objects.get(driver_license_number=self.request.user.username)
        except Offender.DoesNotExist:
            raise PermissionDenied("Citizen profile not found.")
            
        if booking.offender != offender:
            raise PermissionDenied("You can only dispute your own tickets.")
            
        if booking.payment_status == 'Paid':
            raise PermissionDenied("This citation has already been Paid and cleared.")
            
        serializer.save(offender=offender, status='Pending')


class AdminTicketDisputeViewSet(viewsets.ModelViewSet):
    serializer_class = TicketDisputeSerializer
    permission_classes = [IsAuthenticated]
    queryset = TicketDispute.objects.all().order_by('-submitted_at')

    def get_queryset(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only administrators can view all citizen disputes.")
        return super().get_queryset()

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only administrators can review citizen disputes.")
            
        dispute = self.get_object()
        review_status = request.data.get('status')
        review_comments = request.data.get('review_comments', '')

        if review_status not in ['Approved', 'Rejected']:
            return Response({'error': 'Invalid status. Choose Approved or Rejected.'}, status=status.HTTP_400_BAD_REQUEST)

        dispute.status = review_status
        dispute.review_comments = review_comments
        dispute.reviewed_at = timezone.now()
        dispute.save()

        # If Approved, cancel the fine booking
        if review_status == 'Approved':
            booking = dispute.booking
            booking.payment_status = 'Cancelled'
            booking.save()

        return Response(TicketDisputeSerializer(dispute).data)


class CitizenProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = CitizenProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        try:
            return Offender.objects.get(driver_license_number=self.request.user.username)
        except Offender.DoesNotExist:
            raise PermissionDenied("Citizen profile not found.")

    def perform_update(self, serializer):
        # Save profile
        offender = serializer.save()
        
        # Sync email to the authentication user
        user = self.request.user
        email = serializer.validated_data.get('email')
        if email:
            user.email = email
            user.save(update_fields=['email'])

        # Sync phone number to the authentication user as well
        phone_number = serializer.validated_data.get('phone_number')
        if phone_number is not None:
            user.phone_number = phone_number
            user.save(update_fields=['phone_number'])

        # Sync registry record
        from .models import DriverInformation
        DriverInformation.objects.filter(license_number=offender.driver_license_number).update(
            phone_number=offender.phone_number or '',
            email=offender.email or ''
        )


class CitizenChangePasswordView(generics.GenericAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user = request.user
        if not user.check_password(serializer.validated_data['old_password']):
            return Response({'old_password': ['Incorrect password.']}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(serializer.validated_data['new_password'])
        user.save()
        return Response({'message': 'Password successfully updated.'}, status=status.HTTP_200_OK)


class AdminOfficerBulkImportView(APIView):
    permission_classes = [IsAuthenticated]
    from rest_framework.parsers import MultiPartParser, FormParser
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        import csv
        import openpyxl
        import secrets
        import string
        from io import StringIO, BytesIO
        from django.contrib.auth import get_user_model
        from django.core.mail import send_mail
        from django.conf import settings
        from django.core.exceptions import PermissionDenied
        from django.db import transaction

        # Enforce strict RBAC: Only superusers/admins can access this
        if not request.user.is_superuser:
            raise PermissionDenied("Security Alert: Only administrators can bulk import officers.")

        file = request.data.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Handle file names (in tests/scripts, file might be passed as a string or a file object)
        if isinstance(file, str):
            file_name = 'officers.csv'
            file_content = file
        else:
            file_name = file.name.lower()
            file_content = file.read()

        rows = []

        try:
            if file_name.endswith('.csv'):
                if isinstance(file_content, bytes):
                    decoded_file = file_content.decode('utf-8')
                else:
                    decoded_file = file_content
                csv_data = csv.DictReader(StringIO(decoded_file))
                for row in csv_data:
                    def get_val_csv(keys, default=''):
                        for k in keys:
                            for row_k, row_v in row.items():
                                if row_k and row_k.strip().lower() == k.lower():
                                    return str(row_v).strip() if row_v is not None else default
                        return default
                    
                    staff_id = get_val_csv(['Staff ID', 'username', 'Staff ID (Optional)', 'Staff ID / Username', 'badge_number'])
                    first_name = get_val_csv(['First Name', 'Name', 'Full Name', 'Firstname'])
                    last_name = get_val_csv(['Last Name', 'Lastname'])
                    rank = get_val_csv(['Rank'], 'Officer')
                    email = get_val_csv(['Email', 'Email Address', 'email'])
                    
                    rows.append({
                        'staff_id': staff_id,
                        'first_name': first_name,
                        'last_name': last_name,
                        'rank': rank,
                        'email': email
                    })
            elif file_name.endswith('.xlsx'):
                if isinstance(file_content, str):
                    file_bytes = file_content.encode('utf-8')
                else:
                    file_bytes = file_content
                wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                sheet = wb.active
                
                # Extract headers
                headers = []
                for cell in next(sheet.iter_rows(max_row=1)):
                    headers.append(str(cell.value).strip() if cell.value is not None else '')

                # Read data rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Skip completely empty rows
                    if not any(row):
                        continue
                    
                    row_data = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_data[header] = row[idx]
                    
                    def get_val_case_insensitive(keys, default=''):
                        for k in keys:
                            for row_k, row_v in row_data.items():
                                if row_k.strip().lower() == k.lower():
                                    return str(row_v).strip() if row_v is not None else default
                        return default

                    staff_id = get_val_case_insensitive(['Staff ID', 'username', 'Staff ID (Optional)', 'Staff ID / Username', 'badge_number'])
                    first_name = get_val_case_insensitive(['First Name', 'Name', 'Full Name', 'Firstname'])
                    last_name = get_val_case_insensitive(['Last Name', 'Lastname'])
                    rank = get_val_case_insensitive(['Rank'], 'Officer')
                    email = get_val_case_insensitive(['Email', 'Email Address', 'email'])
                    
                    rows.append({
                        'staff_id': staff_id,
                        'first_name': first_name,
                        'last_name': last_name,
                        'rank': rank,
                        'email': email
                    })
            elif file_name.endswith('.xls'):
                import xlrd
                if isinstance(file_content, str):
                    file_bytes = file_content.encode('utf-8')
                else:
                    file_bytes = file_content
                wb = xlrd.open_workbook(file_contents=file_bytes)
                sheet = wb.sheet_by_index(0)
                
                headers = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(0, col_idx)
                    headers.append(str(val).strip() if val is not None else '')
                
                for row_idx in range(1, sheet.nrows):
                    row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                    if not any(str(r).strip() for r in row_values):
                        continue
                    
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row_values):
                            row_data[header] = row_values[col_idx]
                            
                    def get_val_case_insensitive(keys, default=''):
                        for k in keys:
                            for row_k, row_v in row_data.items():
                                if row_k.strip().lower() == k.lower():
                                    if isinstance(row_v, float) and row_v.is_integer():
                                        row_v = int(row_v)
                                    return str(row_v).strip() if row_v is not None else default
                        return default

                    staff_id = get_val_case_insensitive(['Staff ID', 'username', 'Staff ID (Optional)', 'Staff ID / Username', 'badge_number'])
                    first_name = get_val_case_insensitive(['First Name', 'Name', 'Full Name', 'Firstname'])
                    last_name = get_val_case_insensitive(['Last Name', 'Lastname'])
                    rank = get_val_case_insensitive(['Rank'], 'Officer')
                    email = get_val_case_insensitive(['Email', 'Email Address', 'email'])
                    
                    rows.append({
                        'staff_id': staff_id,
                        'first_name': first_name,
                        'last_name': last_name,
                        'rank': rank,
                        'email': email
                    })
            else:
                return Response({'error': 'Unsupported file format. Please upload an Excel (.xlsx/.xls) or CSV (.csv) file.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Failed to read file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        User = get_user_model()
        created_count = 0
        errors = []

        for idx, row in enumerate(rows, start=2): # assuming header is row 1
            staff_id = row['staff_id'].strip()
            first_name = row['first_name'].strip()
            last_name = row['last_name'].strip()
            rank = row['rank'].strip() or 'Officer'
            email = row['email'].strip()

            # Validation
            if not first_name:
                errors.append({'row': idx, 'error': 'First Name / Name is required.'})
                continue
            if not email:
                errors.append({'row': idx, 'error': 'Email is required.'})
                continue

            # Generate Username (Staff ID) if not provided
            if not staff_id:
                is_unique = False
                while not is_unique:
                    rand_digits = ''.join(secrets.choice(string.digits) for _ in range(4))
                    generated_id = f"OYR{rand_digits}"
                    if not User.objects.filter(username=generated_id).exists():
                        staff_id = generated_id
                        is_unique = True

            # Double check username
            if User.objects.filter(username=staff_id).exists():
                errors.append({'row': idx, 'error': f"Staff ID '{staff_id}' is already registered."})
                continue

            # Double check email
            if User.objects.filter(email=email).exists():
                errors.append({'row': idx, 'error': f"Email '{email}' is already registered."})
                continue

            # Generate password
            alphabet = string.ascii_letters + string.digits
            password = ''.join(secrets.choice(alphabet) for _ in range(8))

            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=staff_id,
                        email=email,
                        password=password,
                        first_name=f"{rank} {first_name}",
                        last_name=last_name
                    )
                    user.is_staff = True
                    user.is_officer = True
                    user.is_active = True
                    user.save()

                    # Send Email
                    subject = "Your OYRTMA Officer Credentials"
                    message = (
                        f"Hello {rank} {first_name} {last_name},\n\n"
                        f"An official OYRTMA Field Officer account has been created for you by the System Administrator via bulk import.\n\n"
                        f"Your account details are:\n"
                        f"Staff ID / Username: {staff_id}\n"
                        f"Password: {password}\n\n"
                        f"You can log into the Command Center and the Field Ticket App using the following link:\n"
                        f"http://localhost:5174/ \n\n"
                        f"Please make sure to log in and change your password to keep your credentials secure.\n\n"
                        f"Stay safe on the roads!\n"
                        f"- OYRTMA System Administrator"
                    )
                    
                    try:
                        send_mail(
                            subject,
                            message,
                            settings.EMAIL_HOST_USER,
                            [email],
                            fail_silently=False
                        )
                    except Exception:
                        pass
                
                created_count += 1
            except Exception as e:
                errors.append({'row': idx, 'error': f"Failed to create user: {str(e)}"})

        return Response({
            'message': f'Bulk import complete. {created_count} officer(s) created successfully.',
            'created_count': created_count,
            'error_count': len(errors),
            'errors': errors
        }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)

